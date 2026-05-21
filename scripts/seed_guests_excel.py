import os
import psycopg2
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

# Database Connection Details
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")

EXCEL_FILE_PATH = "Guests.xlsx"

def parse_boolean(val):
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    str_val = str(val).strip().lower()
    return str_val in ("true", "1", "yes", "y")

def parse_int(val, default=None):
    if val is None or str(val).strip() == "":
        return default
    try:
        return int(float(val))
    except ValueError:
        return default

def main():
    if not os.path.exists(EXCEL_FILE_PATH):
        print(f"Error: Excel file '{EXCEL_FILE_PATH}' not found.")
        return

    try:
        conn = psycopg2.connect(
            host=DB_HOST,
            port=DB_PORT,
            database=DB_NAME,
            user=DB_USER,
            password=DB_PASSWORD
        )
        cur = conn.cursor()
        print("Connected to the database successfully.")
    except Exception as e:
        print(f"Database connection error: {e}")
        return

    try:
        wb = load_workbook(filename=EXCEL_FILE_PATH, data_only=True)
        sheet = wb.active
        print(f"Reading sheet: '{sheet.title}'...")

        current_group_name = None
        current_group_id = None
        inserted_guests_map = {}
        group_cache = {}

        # Expected number of columns to extract (0 to 6 = 7 columns)
        EXPECTED_COLS = 6

        for row_idx, raw_row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
            # FIX: Pad the tuple with None values if the row is shorter than expected
            row = raw_row + (None,) * (EXPECTED_COLS - len(raw_row)) if len(raw_row) < EXPECTED_COLS else raw_row

            excel_group_name  = row[0]
            guest_name        = row[1]
            song_requests_raw = row[2]
            plus_one_raw      = row[3]
            dependents_raw    = row[4]
            after_party_raw   = row[5]
            # ignore email
            
            # 1. Update Group Context
            if excel_group_name is not None and str(excel_group_name).strip() != "":
                current_group_name = str(excel_group_name).strip()
                
                if current_group_name not in group_cache:
                    cur.execute("SELECT id FROM groups WHERE group_name = %s;", (current_group_name,))
                    group_row = cur.fetchone()
                    
                    if group_row:
                        current_group_id = group_row[0]
                    else:
                        cur.execute(
                            "INSERT INTO groups (group_name) VALUES (%s) RETURNING id;", 
                            (current_group_name,)
                        )
                        current_group_id = cur.fetchone()[0]
                    
                    group_cache[current_group_name] = current_group_id
                else:
                    current_group_id = group_cache[current_group_name]

            # Skip truly empty rows
            if guest_name is None or str(guest_name).strip() == "":
                continue

            if current_group_id is None:
                print(f"Warning: Row {row_idx} has guest '{guest_name}' but no group context. Skipping.")
                continue

            # 2. Normalize guest parameters
            cleaned_guest_name = str(guest_name).strip()
            plus_one_allowed = parse_boolean(plus_one_raw)
            has_dependents = parse_boolean(dependents_raw)
            song_requests = parse_int(song_requests_raw, default=2)
            after_party = parse_boolean(after_party_raw)

            # 3. Insert into the guests table
            insert_query = """
                INSERT INTO guests (
                    group_id, 
                    name, 
                    plus_one_allowed, 
                    has_dependents, 
                    song_requests, 
                    after_party
                ) VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING guest_id;
            """
            
            cur.execute(insert_query, (
                current_group_id,
                cleaned_guest_name,
                plus_one_allowed,
                has_dependents,
                song_requests,
                after_party
            ))
            
            new_guest_id = cur.fetchone()[0]
            inserted_guests_map[cleaned_guest_name] = new_guest_id

        conn.commit()
        print("Success! Excel file parsed and database seeded successfully.")

    except Exception as e:
        conn.rollback()
        # Enhanced error feedback so you know exactly where it broke
        print(f"\nAn error occurred on row {row_idx if 'row_idx' in locals() else 'unknown'}.")
        print(f"Changes have been rolled back.\nDetails: {e}")
    finally:
        cur.close()
        conn.close()
        print("Database connection closed.")

if __name__ == "__main__":
    main()